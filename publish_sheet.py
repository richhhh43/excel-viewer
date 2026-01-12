#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# =========================
# CONFIG
# =========================
EXCEL_PATH = r"C:\Users\rich_\ncaa 1.xlsm"
SHEET_NAME = "Edges"          # must match exactly (case-sensitive)
OUT_CSV = Path("data/latest.csv")

MAX_ROWS = 8000               # cap so it never runs forever
LAST_COL = 12                 # A..L (includes QR Code if you have it already; safe either way)

def s(x):
    return "" if x is None else str(x).strip()

def is_numberish(x) -> bool:
    if x is None:
        return False
    # accept ints/floats, or strings like "484"
    try:
        float(str(x).strip().replace(",", ""))
        return True
    except Exception:
        return False

def find_header_row(ws, max_scan=25):
    for r in range(1, max_scan + 1):
        v = ws[f"A{r}"].value
        if isinstance(v, str):
            t = v.strip().lower().replace(" ", "")
            if t in ("event#", "event"):
                return r
    return 1

def make_unique(headers):
    seen = {}
    out = []
    for h in headers:
        base = h if h else "Col"
        if base not in seen:
            seen[base] = 0
            out.append(base)
        else:
            seen[base] += 1
            out.append(f"{base}.{seen[base]}")
    return out

def fmt_pct(x):
    """0.2804 -> 28.04%  |  28.04 -> 28.04%  |  '' stays '' """
    if x is None or x == "":
        return ""
    try:
        v = float(str(x).strip().replace("%", "").replace(",", ""))
    except Exception:
        return s(x)
    if 0 <= v <= 1:
        v *= 100.0
    return f"{v:.2f}%"

def fmt_money(x):
    """100 / $100 / 100.0 -> $100  (no decimals)"""
    if x is None or x == "":
        return ""
    txt = str(x).strip()
    txt = txt.replace("$", "").replace(",", "")
    try:
        v = float(txt)
        return f"${int(round(v))}"
    except Exception:
        # if it's already something like "$100"
        return s(x) if s(x).startswith("$") else f"${s(x)}"

def main():
    t0 = time.time()
    print("[publish] starting...")
    print(f"[publish] excel: {EXCEL_PATH}")
    print(f"[publish] sheet: {SHEET_NAME}")

    print("[publish] loading workbook (read_only)...")
    wb = load_workbook(EXCEL_PATH, data_only=False, keep_vba=True, read_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'Sheet "{SHEET_NAME}" not found. Sheets: {wb.sheetnames}')

    ws = wb[SHEET_NAME]
    header_row = find_header_row(ws)
    print(f"[publish] header_row={header_row}")

    # Read headers A..L (or whatever you actually have)
    headers = []
    for c in range(1, LAST_COL + 1):
        headers.append(s(ws.cell(row=header_row, column=c).value) or f"Col{c}")
    cols = make_unique(headers)
    print(f"[publish] columns: {cols}")

    # Bulk read rows
    rows = []
    start = header_row + 1

    for i, rowvals in enumerate(
        ws.iter_rows(
            min_row=start,
            max_row=start + MAX_ROWS - 1,
            min_col=1,
            max_col=LAST_COL,
            values_only=True
        ),
        start=1
    ):
        evt = rowvals[0]

        # Stop at first blank
        if evt in (None, ""):
            print(f"[publish] stop at row {start + i - 1} (blank Event #)")
            break

        # Skip junk/non-data rows (fixes the <openpyxl.work... issue)
        if not is_numberish(evt):
            continue

        rows.append({cols[c]: rowvals[c] for c in range(LAST_COL)})

        if i % 1000 == 0:
            print(f"[publish] scanned {i} rows... kept {len(rows)}")

    df = pd.DataFrame(rows)
    print(f"[publish] kept rows: {len(df)}")

    # Identify columns we care about
    wager_col = "Wager" if "Wager" in df.columns else None
    win_col = "Win%" if "Win%" in df.columns else None
    edge_col = "Edge" if "Edge" in df.columns else None

    # Pick column: prefer Market.1 if present else last Market*
    market_cols = [c for c in df.columns if c.lower().startswith("market")]
    pick_col = "Market.1" if "Market.1" in df.columns else (market_cols[-1] if market_cols else "")

    # Format columns for display on the website
    if win_col:
        df[win_col] = df[win_col].apply(fmt_pct)
    if edge_col:
        df[edge_col] = df[edge_col].apply(fmt_pct)
    if wager_col:
        df[wager_col] = df[wager_col].apply(fmt_money)

    # Build QR Code text (ALWAYS uses formatted values)
    def build_qr(r):
        evt = s(r.get("Event #"))
        dt  = s(r.get("Date and Time"))
        vis = s(r.get("Visitor"))
        home= s(r.get("Home"))
        mkt = s(r.get("Market"))
        odds= s(r.get("Odds"))
        prop= s(r.get("Prop"))
        win = s(r.get("Win%"))   # already formatted like 28.04%
        edge= s(r.get("Edge"))   # already formatted like 12.17%
        pick= s(r.get(pick_col)) if pick_col else ""
        wager = s(r.get("Wager"))  # already like $100

        return (
            f"ALC|EVT:{evt}"
            f"|DT:{dt}"
            f"|V:{vis}"
            f"|H:{home}"
            f"|M:{mkt}"
            f"|OD:{odds}"
            f"|P:{prop}"
            f"|WIN:{win}"
            f"|EDGE:{edge}"
            f"|PICK:{pick}"
            f"|WAGER:{wager}"
        )

    df["QR Code"] = df.apply(lambda row: build_qr(row.to_dict()), axis=1)

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUT_CSV, index=False)

    print(f"[publish] saved: {OUT_CSV.as_posix()}")
    print(f"[publish] done in {time.time() - t0:.2f}s")

if __name__ == "__main__":
    main()
